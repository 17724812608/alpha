import requests
from openpyxl import load_workbook


def read_case(filename,sheetname):   #定义函数
    wb = load_workbook(filename)     #打开Excel文件
    sh = wb[sheetname]               #打开某个表单
    rows = sh.max_row                 #获取最大行数
    case_list = []                   #空列表
    for i in range(2,rows+1):         #循环获取
        case_dict = dict(             #用字典存储
        case_id = (sh.cell(row=i,column=1)).value,      #获取case_id
        url = sh.cell(row=i,column=5).value,            #获取url
        data = sh.cell(row=i,column=6).value,           #获取data
        expected = sh.cell(row=i,column=7).value,       #获取期望
        )
        case_list.append(case_dict)    #每次循环都把生成的dict追加到list中
    return case_list                    #设置返回值

# #封装写入测试结果
def write_result(filename,sheetname,row,column,final_result):
    wb = load_workbook(filename)   #打开Excel文件
    sh = wb[sheetname]            #打开某个表单
    rows = sh.max_row             #获取最大行数
    sh.cell(row=row,column=column).value = final_result    #写入数据
    wb.save(filename)             #保存

#发送接口需求
def api_fun(url,data):
    header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url,json=data,headers=header).json()
    return res        #设置返回值